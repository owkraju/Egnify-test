import openpyxl
b=['Teacher wise class timetable - Hindi.xlsx','Teacher wise class timetable - Kannada.xlsx','Teacher wise class timetable - Maths.xlsx','Teacher wise class timetable - Science.xlsx','Teacher wise class timetable - English.xlsx']
time_class_wise={}
timetable=[]
classes={}
subjects={}
timings={}
for k in range(0,5):
	book = openpyxl.load_workbook(b[k])
	sub_name=b[k].split('-')
	sub_name=sub_name[1].split('.')
	
	sheet = book.active
	clas='6th'
	i=1	
	
	for c in range(6,10+1):
		while(i<=7):
			for j in range(1,sheet.max_row+1):
				cl=sheet.cell(row=j,column=i).value
				if (cl==str(c)+'th'):	
					
						d=str(sheet.cell(row=1,column=i).value)
						t=str(sheet.cell(row=j,column=1).value)
						timings[d]=t
							
				subjects[sub_name[0]]=timings
					

			i=i+1
			
		classes[str(c)+'th']=subjects
	time_class_wise[0]=classes
for i in range(6,11):
	print "class name:{}".format(i)
	print time_class_wise[0][str(i)+'th']
	print '\n'
